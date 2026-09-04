import tempfile
import csv
from io import StringIO, BytesIO
from datetime import timedelta

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from dashboard.models import DashboardUserProfile
from dashboard.views import EXPORT_COLUMNS
from facilities.forms import FacilityForm
from facilities.models import Facility
from feedback.models import Feedback, RatingResponse

User = get_user_model()
TEST_MEDIA_ROOT = tempfile.mkdtemp()


@override_settings(
    MEDIA_ROOT=TEST_MEDIA_ROOT,
    STATICFILES_STORAGE="django.contrib.staticfiles.storage.StaticFilesStorage",
)
class DashboardAccessTests(TestCase):
    def _create_submission(self, facility, *, rating_pairs, **shared_fields):
        submission = Feedback.objects.create(
            facility=facility,
            **shared_fields,
        )
        RatingResponse.objects.bulk_create(
            [
                RatingResponse(
                    submission=submission,
                    category=category,
                    rating=rating,
                    comment=comment,
                )
                for category, rating, comment in rating_pairs
            ]
        )
        return submission

    def setUp(self):
        self.facility_a = Facility.objects.create(name="Facility A", district="A", province="P1")
        self.facility_b = Facility.objects.create(name="Facility B", district="B", province="P2")
        self.dashboard_user = User.objects.create_user(username="dash", password="secret123")
        profile = DashboardUserProfile.objects.get(user=self.dashboard_user)
        profile.is_dashboard_user = True
        profile.facility = self.facility_a
        profile.save()

        self._create_submission(
            self.facility_a,
            rating_pairs=[(Feedback.Category.WAITING_TIME, 4, "Good wait time")],
            gender=Feedback.Gender.FEMALE,
        )
        self._create_submission(
            self.facility_b,
            rating_pairs=[(Feedback.Category.CLEANLINESS, 2, "Needs improvement")],
            gender=Feedback.Gender.MALE,
        )

    def test_dashboard_user_only_sees_assigned_facility_feedback(self):
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:feedback_list"))

        returned_facilities = {
            entry.facility.name for entry in response.context["feedback_entries"]
        }
        self.assertEqual(returned_facilities, {"Facility A"})

    def test_feedback_list_paginates_to_ten_responses(self):
        for _ in range(11):
            self._create_submission(
                self.facility_a,
                rating_pairs=[(Feedback.Category.WAITING_TIME, 5, "")],
                gender=Feedback.Gender.FEMALE,
            )

        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:feedback_list"))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["is_paginated"])
        self.assertEqual(response.context["paginator"].per_page, 10)
        self.assertEqual(len(response.context["feedback_entries"]), 10)

    def test_feedback_list_orders_latest_submissions_first(self):
        older_submission = self._create_submission(
            self.facility_a,
            rating_pairs=[(Feedback.Category.WAITING_TIME, 2, "Older")],
            gender=Feedback.Gender.FEMALE,
        )
        latest_submission = self._create_submission(
            self.facility_a,
            rating_pairs=[(Feedback.Category.CLEANLINESS, 5, "Latest")],
            gender=Feedback.Gender.MALE,
        )
        Feedback.objects.filter(pk=older_submission.pk).update(created_at=timezone.now() - timedelta(days=2))
        Feedback.objects.filter(pk=latest_submission.pk).update(created_at=timezone.now())
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:feedback_list"))

        entries = list(response.context["feedback_entries"])
        self.assertGreaterEqual(len(entries), 2)
        self.assertEqual(entries[0].pk, latest_submission.pk)
        self.assertEqual(entries[-1].pk, older_submission.pk)

    def test_feedback_list_filters_by_suggested_change(self):
        matching_submission = self._create_submission(
            self.facility_a,
            rating_pairs=[(Feedback.Category.WAITING_TIME, 4, "")],
            change=Feedback.CHANGE.MORE_MEDICINES,
        )
        self._create_submission(
            self.facility_a,
            rating_pairs=[(Feedback.Category.WAITING_TIME, 4, "")],
            change=Feedback.CHANGE.MORE_WORKERS,
        )
        self.client.login(username="dash", password="secret123")

        response = self.client.get(
            reverse("dashboard:feedback_list"),
            {"change": Feedback.CHANGE.MORE_MEDICINES},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [entry.pk for entry in response.context["feedback_entries"]],
            [matching_submission.pk],
        )

    def test_dashboard_user_export_is_scoped_to_assigned_facility(self):
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:export_csv"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("Facility A", content)
        self.assertNotIn("Facility B", content)

    def test_dashboard_export_includes_extended_feedback_fields(self):
        Feedback.objects.filter(facility=self.facility_a).update(
            age_group=Feedback.AgeGroup.AGE_25_34,
            distance=Feedback.Distance.LESS_THAN_5KM,
            service=Feedback.Service.OTHER,
            service_other="Mental health consultation",
            difficulty=[Feedback.Difficulty.HEARING, Feedback.Difficulty.MOBILITY],
            received_service=Feedback.receivedService.PARTIALLY,
            reason_not_received=Feedback.ReasonNotReceived.MEDICINE,
            reason_not_received_other="",
            referral=Feedback.Referral.YES,
            facility_type=Feedback.FacilityType.HOSPITAL,
            payment=Feedback.Payment.YES,
            insurance=Feedback.INSURANCE.NONE,
            no_insurance_reason=Feedback.NO_INSURANCE_REASON.CASH,
            cash_payment=Feedback.CASH.BETWEEN,
            cash_payment_other="",
            cost=Feedback.COST.YES,
            medicines=Feedback.MEDICINES.NO_PHARMACY,
            revisit=Feedback.REVISIT.NOT_SURE,
            chance=Feedback.CHANCE.NO,
            reason_not_chance=Feedback.REASON_NOT_CHANCE.FAR,
            reason_not_chance_other="",
            change=Feedback.CHANGE.OTHER,
            change_other="Improve triage",
            aob=Feedback.AOB.YES,
            aob_other="Waiting area needs seating",
        )
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:export_csv"))

        rows = list(csv.reader(StringIO(response.content.decode("utf-8"))))
        self.assertEqual(rows[0], [label for label, _getter in EXPORT_COLUMNS])
        self.assertIn("Mental health consultation", rows[1])
        self.assertIn("Hearing (even with hearing aid), Walking or climbing steps", rows[1])
        self.assertIn("Improve triage", rows[1])
        self.assertIn("Waiting area needs seating", rows[1])
        self.assertIn("4", rows[1])

    def test_dashboard_submission_export_has_one_row_per_submission(self):
        self._create_submission(
            self.facility_a,
            rating_pairs=[
                (Feedback.Category.WAITING_TIME, 5, "Very fast"),
                (Feedback.Category.CLEANLINESS, 3, "Average"),
                (Feedback.Category.MEDICATION, 4, "Medicines available"),
            ],
            gender=Feedback.Gender.FEMALE,
            comment="General note",
        )
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:export_csv"))

        rows = list(csv.reader(StringIO(response.content.decode("utf-8"))))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(rows), 3)
        header = rows[0]
        latest_row = next(
            row for row in rows[1:]
            if row[header.index("Availability of Medication comment")] == "Medicines available"
        )
        self.assertIn("Ratings answered", header)
        self.assertIn("Average rating", header)
        self.assertEqual(latest_row[header.index("Ratings answered")], "3")
        self.assertEqual(latest_row[header.index("Waiting time before being seen rating")], "5")
        self.assertEqual(latest_row[header.index("Waiting time before being seen comment")], "Very fast")
        self.assertEqual(latest_row[header.index("Cleanliness of the health facility rating")], "3")
        self.assertEqual(latest_row[header.index("Availability of Medication comment")], "Medicines available")

    def test_dashboard_excel_export_uses_submission_sheet_and_filename(self):
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:export_excel"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("feedback-submissions-export.xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn("Feedback Submissions", workbook.sheetnames)

    def test_dashboard_home_includes_new_analytics_context(self):
        Feedback.objects.filter(facility=self.facility_a).update(
            received_service=Feedback.receivedService.PARTIALLY,
            payment=Feedback.Payment.YES,
            medicines=Feedback.MEDICINES.NO_PHARMACY,
            revisit=Feedback.REVISIT.YES,
            insurance=Feedback.INSURANCE.NHIMA,
            change=Feedback.CHANGE.MORE_MEDICINES,
            reason_not_received=Feedback.ReasonNotReceived.MEDICINE,
        )
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:home"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("patient_experience_cards", response.context)
        self.assertIn("received_service_card", response.context)
        self.assertIn("payment_card", response.context)
        self.assertIn("medicines_card", response.context)
        self.assertIn("revisit_card", response.context)
        self.assertIn("insurance_breakdown", response.context)
        self.assertIn("change_breakdown", response.context)
        self.assertIn("reason_not_received_breakdown", response.context)
        self.assertIn("insurance_chart_summary", response.context)
        self.assertIn("change_chart_summary", response.context)
        self.assertIn("insurance_percentages", response.context)
        self.assertIn("change_percentages", response.context)
        self.assertIn("province_summary", response.context)
        self.assertIn("source_breakdown", response.context)

    def test_dashboard_home_submission_breakdowns_count_parent_submissions_once(self):
        self._create_submission(
            self.facility_a,
            rating_pairs=[
                (Feedback.Category.WAITING_TIME, 5, "Fast"),
                (Feedback.Category.CLEANLINESS, 4, "Clean"),
                (Feedback.Category.MEDICATION, 3, "Available"),
            ],
            submission_source=Feedback.SubmissionSource.QR_PUBLIC,
            insurance=Feedback.INSURANCE.NHIMA,
            change=Feedback.CHANGE.MORE_MEDICINES,
            reason_not_received=Feedback.ReasonNotReceived.MEDICINE,
        )
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:home"))

        self.assertEqual(response.status_code, 200)
        source_breakdown = {item["value"]: item["total"] for item in response.context["source_breakdown"]}
        insurance_breakdown = {item["value"]: item["total"] for item in response.context["insurance_breakdown"]}
        change_breakdown = {item["value"]: item["total"] for item in response.context["change_breakdown"]}
        reason_breakdown = {item["value"]: item["total"] for item in response.context["reason_not_received_breakdown"]}
        facility_breakdown = {item["facility__name"]: item["total"] for item in response.context["facility_breakdown"]}
        province_breakdown = {
            item["facility__province"]: item["total"]
            for item in response.context["province_summary"]["rows"]
        }

        self.assertEqual(source_breakdown[Feedback.SubmissionSource.QR_PUBLIC], 2)
        self.assertEqual(insurance_breakdown[Feedback.INSURANCE.NHIMA], 1)
        self.assertEqual(change_breakdown[Feedback.CHANGE.MORE_MEDICINES], 1)
        self.assertEqual(reason_breakdown[Feedback.ReasonNotReceived.MEDICINE], 1)
        self.assertEqual(facility_breakdown[self.facility_a.name], 2)
        self.assertEqual(province_breakdown[self.facility_a.province], 2)
        self.assertEqual(response.context["public_qr_total"] + response.context["assisted_total"], response.context["total_submissions"])

    def test_dashboard_home_chart_summaries_use_answered_submissions_and_clear_labels(self):
        Feedback.objects.filter(facility=self.facility_a).update(
            insurance=Feedback.INSURANCE.NONE,
            change=Feedback.CHANGE.STAFF_ATTITUDE,
        )
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:home"))

        self.assertEqual(response.status_code, 200)
        insurance_breakdown = response.context["insurance_breakdown"]
        change_breakdown = response.context["change_breakdown"]
        self.assertEqual(response.context["insurance_chart_summary"], "1 respondents")
        self.assertEqual(response.context["change_chart_summary"], "1 respondents")
        self.assertEqual(response.context["change_chart_note"], "")
        self.assertEqual(insurance_breakdown[0]["label"], "No insurance used")
        self.assertEqual(insurance_breakdown[0]["percentage"], 100.0)
        self.assertEqual(change_breakdown[0]["percentage"], 100.0)

    def test_dashboard_home_source_filter_limits_all_metrics(self):
        self._create_submission(
            self.facility_a,
            rating_pairs=[(Feedback.Category.MEDICATION, 3, "Assisted")],
            submission_source=Feedback.SubmissionSource.ASSISTED_CAPTURE,
            insurance=Feedback.INSURANCE.PRIVATE,
            change=Feedback.CHANGE.OTHER,
            received_service=Feedback.receivedService.YES,
        )
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:home"), {"source": Feedback.SubmissionSource.ASSISTED_CAPTURE})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_submissions"], 1)
        self.assertEqual(response.context["public_qr_total"], 0)
        self.assertEqual(response.context["assisted_total"], 1)
        self.assertEqual(response.context["source_breakdown"][1]["percentage"], 100.0)
        self.assertEqual(response.context["insurance_breakdown"][0]["value"], Feedback.INSURANCE.PRIVATE)

    def test_dashboard_home_period_filter_excludes_older_feedback(self):
        older_submission = self._create_submission(
            self.facility_a,
            rating_pairs=[(Feedback.Category.MEDICATION, 2, "Old")],
            submission_source=Feedback.SubmissionSource.QR_PUBLIC,
        )
        Feedback.objects.filter(pk=older_submission.pk).update(created_at=timezone.now() - timedelta(days=45))
        self.client.login(username="dash", password="secret123")

        response = self.client.get(reverse("dashboard:home"), {"period": "30d"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_submissions"], 1)
        self.assertEqual(response.context["trend_totals"], [1])

    def test_dashboard_home_facility_filter_and_empty_state(self):
        self.client.login(username="dash", password="secret123")

        response = self.client.get(
            reverse("dashboard:home"),
            {"period": "today", "source": Feedback.SubmissionSource.ASSISTED_CAPTURE},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_submissions"], 0)
        self.assertContains(response, "No feedback available for the selected period.")

    def test_dashboard_user_can_open_detail_for_assigned_facility_feedback_only(self):
        visible_feedback = Feedback.objects.filter(facility=self.facility_a).first()
        hidden_feedback = Feedback.objects.filter(facility=self.facility_b).first()
        self.client.login(username="dash", password="secret123")

        visible_response = self.client.get(
            reverse("dashboard:feedback_detail", args=[visible_feedback.pk])
        )
        hidden_response = self.client.get(
            reverse("dashboard:feedback_detail", args=[hidden_feedback.pk])
        )

        self.assertEqual(visible_response.status_code, 200)
        self.assertEqual(visible_response.context["entry"].facility, self.facility_a)
        self.assertEqual(hidden_response.status_code, 404)


@override_settings(
    MEDIA_ROOT=TEST_MEDIA_ROOT,
    STATICFILES_STORAGE="django.contrib.staticfiles.storage.StaticFilesStorage",
)
class FacilityManagementTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user(
            username="staff",
            password="secret123",
            is_staff=True,
        )
        self.facility = Facility.objects.create(
            name="Delete Me Clinic",
            district="Lusaka",
            province="Lusaka",
        )

    def test_staff_user_can_delete_facility_from_dashboard(self):
        self.client.login(username="staff", password="secret123")

        response = self.client.post(reverse("dashboard:facility_delete", args=[self.facility.pk]))

        self.assertRedirects(response, reverse("dashboard:facility_list"))
        self.assertFalse(Facility.objects.filter(pk=self.facility.pk).exists())

    def test_facility_form_rejects_district_not_in_selected_province(self):
        form = FacilityForm(
            data={
                "name": "Mismatch Clinic",
                "province": "Lusaka",
                "district": "Kitwe",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("district", form.errors)

    def test_facility_list_paginates_to_ten_entries(self):
        for index in range(12):
            Facility.objects.create(
                name=f"Facility {index}",
                district="Lusaka",
                province="Lusaka",
            )
        self.client.login(username="staff", password="secret123")

        response = self.client.get(reverse("dashboard:facility_list"))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["is_paginated"])
        self.assertEqual(response.context["paginator"].per_page, 10)
        self.assertEqual(len(response.context["facilities"]), 10)

    def test_facility_list_search_filters_by_name_district_or_province(self):
        Facility.objects.create(
            name="Kanyama Level One Hospital",
            district="Lusaka",
            province="Lusaka",
        )
        Facility.objects.create(
            name="Mansa General Hospital",
            district="Mansa",
            province="Luapula",
        )
        self.client.login(username="staff", password="secret123")

        response = self.client.get(reverse("dashboard:facility_list"), {"search": "Luapula"})

        self.assertEqual(response.status_code, 200)
        returned_names = {facility.name for facility in response.context["facilities"]}
        self.assertEqual(returned_names, {"Mansa General Hospital"})


@override_settings(
    MEDIA_ROOT=TEST_MEDIA_ROOT,
    STATICFILES_STORAGE="django.contrib.staticfiles.storage.StaticFilesStorage",
)
class DashboardUserManagementTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user(
            username="staff-admin",
            password="secret123",
            is_staff=True,
        )
        self.facility = Facility.objects.create(
            name="Kabwata General Hospital",
            district="Lusaka",
            province="Lusaka",
        )
        self.dashboard_user = User.objects.create_user(
            username="field-user",
            password="oldpassword123",
        )

    def test_staff_user_can_open_edit_user_view(self):
        self.client.login(username="staff-admin", password="secret123")

        response = self.client.get(reverse("dashboard:user_update", args=[self.dashboard_user.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["target_user"], self.dashboard_user)
        self.assertContains(response, "Reset password")
        self.assertContains(response, "Delete user")

    def test_user_list_paginates_to_ten_entries(self):
        for index in range(12):
            User.objects.create_user(
                username=f"user-{index}",
                password="secret123",
            )
        self.client.login(username="staff-admin", password="secret123")

        response = self.client.get(reverse("dashboard:user_list"))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["is_paginated"])
        self.assertEqual(response.context["paginator"].per_page, 10)
        self.assertEqual(len(response.context["users"]), 10)

    def test_user_list_search_filters_by_username_name_or_email(self):
        self.dashboard_user.first_name = "Kabwata"
        self.dashboard_user.last_name = "Officer"
        self.dashboard_user.email = "kabwata@example.com"
        self.dashboard_user.save()
        User.objects.create_user(
            username="another-user",
            email="another@example.com",
            password="secret123",
        )
        self.client.login(username="staff-admin", password="secret123")

        response = self.client.get(reverse("dashboard:user_list"), {"search": "kabwata"})

        self.assertEqual(response.status_code, 200)
        returned_usernames = {user.username for user in response.context["users"]}
        self.assertEqual(returned_usernames, {"field-user"})

    def test_staff_user_can_update_user_details_from_edit_view(self):
        self.client.login(username="staff-admin", password="secret123")

        response = self.client.post(
            reverse("dashboard:user_update", args=[self.dashboard_user.pk]),
            data={
                "username": "field-user-edited",
                "first_name": "Kabwata",
                "last_name": "Officer",
                "email": "kabwata@example.com",
                "is_dashboard_user": "on",
                "facility": str(self.facility.pk),
            },
        )

        self.assertRedirects(response, reverse("dashboard:user_list"))
        self.dashboard_user.refresh_from_db()
        profile = DashboardUserProfile.objects.get(user=self.dashboard_user)
        self.assertEqual(self.dashboard_user.username, "field-user-edited")
        self.assertEqual(self.dashboard_user.first_name, "Kabwata")
        self.assertEqual(self.dashboard_user.last_name, "Officer")
        self.assertEqual(self.dashboard_user.email, "kabwata@example.com")
        self.assertTrue(profile.is_dashboard_user)
        self.assertEqual(profile.facility, self.facility)

    def test_staff_user_can_reset_dashboard_user_password(self):
        self.client.login(username="staff-admin", password="secret123")

        response = self.client.post(
            reverse("dashboard:user_password_reset", args=[self.dashboard_user.pk]),
            data={
                "new_password1": "newsecurepass456",
                "new_password2": "newsecurepass456",
            },
        )

        self.assertRedirects(response, reverse("dashboard:user_list"))
        self.dashboard_user.refresh_from_db()
        self.assertTrue(self.dashboard_user.check_password("newsecurepass456"))

    def test_non_staff_user_cannot_access_password_reset_view(self):
        self.client.login(username="field-user", password="oldpassword123")

        response = self.client.get(
            reverse("dashboard:user_password_reset", args=[self.dashboard_user.pk])
        )

        self.assertEqual(response.status_code, 403)

    def test_staff_user_can_delete_non_admin_user(self):
        self.client.login(username="staff-admin", password="secret123")

        response = self.client.post(
            reverse("dashboard:user_delete", args=[self.dashboard_user.pk])
        )

        self.assertRedirects(response, reverse("dashboard:user_list"))
        self.assertFalse(User.objects.filter(pk=self.dashboard_user.pk).exists())

    def test_staff_user_cannot_delete_admin_account(self):
        self.client.login(username="staff-admin", password="secret123")

        response = self.client.post(
            reverse("dashboard:user_delete", args=[self.staff_user.pk])
        )

        self.assertRedirects(response, reverse("dashboard:user_list"))
        self.assertTrue(User.objects.filter(pk=self.staff_user.pk).exists())

    def test_dashboard_user_can_update_own_profile_and_contact_information(self):
        self.client.login(username="field-user", password="oldpassword123")

        response = self.client.post(
            reverse("dashboard:account"),
            data={
                "account_action": "profile",
                "username": "field-user-updated",
                "first_name": "Field",
                "last_name": "Officer",
                "email": "field.officer@example.com",
            },
        )

        self.assertRedirects(response, reverse("dashboard:account"))
        self.dashboard_user.refresh_from_db()
        self.assertEqual(self.dashboard_user.username, "field-user-updated")
        self.assertEqual(self.dashboard_user.first_name, "Field")
        self.assertEqual(self.dashboard_user.last_name, "Officer")
        self.assertEqual(self.dashboard_user.email, "field.officer@example.com")

    def test_dashboard_user_can_change_own_password(self):
        self.client.login(username="field-user", password="oldpassword123")

        response = self.client.post(
            reverse("dashboard:account"),
            data={
                "account_action": "password",
                "old_password": "oldpassword123",
                "new_password1": "newstrongpass789",
                "new_password2": "newstrongpass789",
            },
        )

        self.assertRedirects(response, reverse("dashboard:account"))
        self.dashboard_user.refresh_from_db()
        self.assertTrue(self.dashboard_user.check_password("newstrongpass789"))
