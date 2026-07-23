import csv
from datetime import datetime, timedelta
from io import TextIOWrapper

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from facilities.models import Facility
from feedback.audit import log_bulk_event
from feedback.duplicate_detection import (
    DuplicateStatus,
    build_feedback_fingerprint,
    classify_duplicate,
    normalize_text,
)
from feedback.forms import FeedbackForm
from feedback.models import Feedback, ImportBatch
from feedback.submission_service import create_feedback_entries_from_cleaned_data


MAX_IMPORT_FILE_SIZE = int(getattr(settings, "MAX_IMPORT_FILE_SIZE", 5 * 1024 * 1024))
MAX_IMPORT_ROW_COUNT = int(getattr(settings, "MAX_IMPORT_ROW_COUNT", 500))
ALLOW_POSSIBLE_DUPLICATES = (
    str(getattr(settings, "IMPORT_ALLOW_POSSIBLE_DUPLICATES", "False")).lower() == "true"
)

CATEGORY_FIELD_MAP = {
    Feedback.Category.WAITING_TIME: ("rating_waiting_time", "comment_waiting_time"),
    Feedback.Category.STAFF_ATTITUDE: ("rating_staff_attitude", "comment_staff_attitude"),
    Feedback.Category.CLEANLINESS: ("rating_cleanliness", "comment_cleanliness"),
    Feedback.Category.EXPLANATION: ("rating_explanation", "comment_explanation"),
    Feedback.Category.MEDICATION: ("rating_medication", "comment_medication"),
}

FIELD_COLUMNS = [
    "facility_code",
    "submitted_on",
    "gender",
    "age_group",
    "distance",
    "service",
    "service_other",
    "difficulty",
    "received_service",
    "reason_not_received",
    "reason_not_received_other",
    "referral",
    "facility_type",
    "facility_type_other",
    "payment",
    "insurance",
    "no_insurance_reason",
    "no_insurance_reason_other",
    "cash_payment",
    "cash_payment_other",
    "cost",
    "medicines",
    "revisit",
    "chance",
    "reason_not_chance",
    "reason_not_chance_other",
    "change",
    "change_other",
    "aob",
    "aob_other",
]

TEMPLATE_COLUMNS = FIELD_COLUMNS + [
    rating_column
    for rating_column, _comment_column in CATEGORY_FIELD_MAP.values()
] + [
    comment_column
    for _rating_column, comment_column in CATEGORY_FIELD_MAP.values()
]

CHOICE_FIELD_MAP = {
    "gender": Feedback.Gender.choices,
    "age_group": Feedback.AgeGroup.choices,
    "distance": Feedback.Distance.choices,
    "service": Feedback.Service.choices,
    "received_service": Feedback.receivedService.choices,
    "reason_not_received": Feedback.ReasonNotReceived.choices,
    "referral": Feedback.Referral.choices,
    "facility_type": Feedback.FacilityType.choices,
    "payment": Feedback.Payment.choices,
    "insurance": Feedback.INSURANCE.choices,
    "no_insurance_reason": Feedback.NO_INSURANCE_REASON.choices,
    "cash_payment": Feedback.CASH.choices,
    "cost": Feedback.COST.choices,
    "medicines": Feedback.MEDICINES.choices,
    "revisit": Feedback.REVISIT.choices,
    "chance": Feedback.CHANCE.choices,
    "reason_not_chance": Feedback.REASON_NOT_CHANCE.choices,
    "change": Feedback.CHANGE.choices,
    "aob": Feedback.AOB.choices,
}


def _choice_lookup(choices):
    return {str(value).strip().lower(): value for value, _label in choices}


def workbook_template_response():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Feedback Template"
    sheet.append(TEMPLATE_COLUMNS)
    sheet.append(
        [
            "facility-1",
            timezone.localdate().strftime("%d/%m/%Y"),
            Feedback.Gender.FEMALE,
            Feedback.AgeGroup.AGE_25_34,
            Feedback.Distance.LESS_THAN_5KM,
            Feedback.Service.TREATMENT,
            "",
            Feedback.Difficulty.NONE,
            Feedback.receivedService.YES,
            "",
            "",
            Feedback.Referral.NO,
            "",
            "",
            Feedback.Payment.NO,
            Feedback.INSURANCE.NONE,
            Feedback.NO_INSURANCE_REASON.NONE,
            "",
            "",
            "",
            Feedback.COST.NO,
            Feedback.MEDICINES.YES,
            Feedback.REVISIT.YES,
            Feedback.CHANCE.YES,
            "",
            "",
            Feedback.CHANGE.MORE_MEDICINES,
            "",
            Feedback.AOB.NO,
            "",
            4,
            5,
            4,
            5,
            4,
            "Reasonable wait time",
            "Respectful staff",
            "Clean facility",
            "Clear explanations",
            "Medicines were available",
        ]
    )

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Field", "Guidance"])
    instructions.append(["submitted_on", "Use DD/MM/YYYY format. Leave blank to use today's date."])
    instructions.append(["facility_code", "Use the facility slug-id code from the system, or select a facility during upload."])
    instructions.append(["difficulty", "Separate multiple values with | characters. Example: Seeing (even with glasses)|Communicating"])
    instructions.append(["rating_*", "Ratings must be whole numbers from 1 to 5. At least one rating is required per row."])

    return workbook


def _read_rows_from_batch(batch: ImportBatch):
    file_name = batch.stored_file.name.lower()
    if file_name.endswith(".csv"):
        batch.stored_file.open("rb")
        try:
            text_stream = TextIOWrapper(batch.stored_file.file, encoding="utf-8-sig")
            reader = csv.DictReader(text_stream)
            rows = []
            for row in reader:
                rows.append({key: value for key, value in row.items()})
            return rows
        finally:
            batch.stored_file.close()

    batch.stored_file.open("rb")
    try:
        workbook = load_workbook(batch.stored_file, data_only=False)
        worksheet = workbook.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
        rows = []
        for row in worksheet.iter_rows(min_row=2):
            row_dict = {}
            for header, cell in zip(headers, row):
                if getattr(cell, "data_type", "") == "f":
                    row_dict[header] = "__FORMULA__"
                else:
                    row_dict[header] = cell.value
            rows.append(row_dict)
        return rows
    finally:
        batch.stored_file.close()


def _normalize_difficulty(value):
    if isinstance(value, list):
        return value
    if not value:
        return []
    return [item.strip() for item in str(value).split("|") if item.strip()]


def _normalize_submitted_on(value):
    if not value:
        return timezone.localdate()
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value
    for date_format in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(str(value).strip(), date_format)
            return parsed.date()
        except ValueError:
            continue
    raise ValidationError("Use DD/MM/YYYY or YYYY-MM-DD for submitted_on.")


def _resolve_facility(raw_value, default_facility):
    if default_facility:
        return default_facility
    if not raw_value:
        raise ValidationError("facility_code is required when no upload facility is selected.")
    raw_value = normalize_text(raw_value)
    for facility in Facility.objects.all():
        if normalize_text(facility.get_feedback_slug()) == raw_value:
            return facility
    raise ValidationError("Unknown facility code.")


def validate_import_batch(batch: ImportBatch, accessible_facilities):
    if batch.stored_file.size > MAX_IMPORT_FILE_SIZE:
        raise ValidationError(f"File exceeds the maximum size of {MAX_IMPORT_FILE_SIZE} bytes.")

    batch.status = ImportBatch.Status.VALIDATING
    batch.save(update_fields=["status"])

    raw_rows = _read_rows_from_batch(batch)
    if not raw_rows:
        batch.status = ImportBatch.Status.VALIDATION_FAILED
        batch.validation_summary = {"errors": ["The uploaded file is empty."], "rows": []}
        batch.save(update_fields=["status", "validation_summary"])
        return batch

    if len(raw_rows) > MAX_IMPORT_ROW_COUNT:
        raise ValidationError(f"File exceeds the maximum row count of {MAX_IMPORT_ROW_COUNT}.")

    headers = {normalize_text(header) for header in raw_rows[0].keys() if header}
    required_headers = {normalize_text(column) for column in TEMPLATE_COLUMNS}
    missing_headers = sorted(required_headers - headers)
    if missing_headers:
        batch.status = ImportBatch.Status.VALIDATION_FAILED
        batch.validation_summary = {
            "errors": [f"Missing required columns: {', '.join(missing_headers)}"],
            "rows": [],
        }
        batch.save(update_fields=["status", "validation_summary"])
        return batch

    accessible_ids = set(accessible_facilities.values_list("id", flat=True))
    row_results = []
    seen_fingerprints = set()
    valid_rows = invalid_rows = duplicate_rows = 0

    for index, raw_row in enumerate(raw_rows, start=2):
        row_errors = []
        normalized_row = {normalize_text(key): value for key, value in raw_row.items()}

        if any(str(value).startswith(("=", "+", "-", "@")) for value in normalized_row.values() if isinstance(value, str)):
            row_errors.append(
                {
                    "field_name": "__row__",
                    "submitted_value": "",
                    "error_message": "Formula-like values are not allowed in imports.",
                }
            )

        if "__FORMULA__" in normalized_row.values():
            row_errors.append(
                {
                    "field_name": "__row__",
                    "submitted_value": "",
                    "error_message": "Formula cells are not allowed in Excel imports.",
                }
            )

        try:
            facility = _resolve_facility(normalized_row.get("facility_code"), batch.facility)
            if facility.id not in accessible_ids:
                raise ValidationError("You are not allowed to import feedback for this facility.")
        except ValidationError as exc:
            row_errors.append(
                {
                    "field_name": "facility_code",
                    "submitted_value": normalized_row.get("facility_code", ""),
                    "error_message": str(exc),
                }
            )
            facility = None

        try:
            submitted_on = _normalize_submitted_on(normalized_row.get("submitted_on"))
            if submitted_on > timezone.localdate() + timedelta(days=2):
                raise ValidationError("Submitted date is too far in the future.")
        except ValidationError as exc:
            row_errors.append(
                {
                    "field_name": "submitted_on",
                    "submitted_value": normalized_row.get("submitted_on", ""),
                    "error_message": str(exc),
                }
            )
            submitted_on = timezone.localdate()

        cleaned_payload = {}
        for field_name in FIELD_COLUMNS[2:]:
            raw_value = normalized_row.get(field_name)
            if field_name == "difficulty":
                cleaned_payload[field_name] = _normalize_difficulty(raw_value)
                continue

            if field_name in CHOICE_FIELD_MAP and raw_value:
                choice_value = _choice_lookup(CHOICE_FIELD_MAP[field_name]).get(normalize_text(raw_value))
                if not choice_value:
                    row_errors.append(
                        {
                            "field_name": field_name,
                            "submitted_value": raw_value,
                            "error_message": "Invalid controlled value.",
                        }
                    )
                cleaned_payload[field_name] = choice_value or ""
            else:
                cleaned_payload[field_name] = raw_value or ""

        ratings = {}
        comments = {}
        for category, (rating_column, comment_column) in CATEGORY_FIELD_MAP.items():
            raw_rating = normalized_row.get(rating_column)
            if raw_rating in ("", None):
                continue
            try:
                rating_value = int(raw_rating)
                if rating_value < 1 or rating_value > 5:
                    raise ValueError
            except (TypeError, ValueError):
                row_errors.append(
                    {
                        "field_name": rating_column,
                        "submitted_value": raw_rating,
                        "error_message": "Ratings must be whole numbers between 1 and 5.",
                    }
                )
                continue
            ratings[category] = rating_value
            comments[category] = normalized_row.get(comment_column, "") or ""

        if not ratings:
            row_errors.append(
                {
                    "field_name": "__ratings__",
                    "submitted_value": "",
                    "error_message": "At least one category rating is required.",
                }
            )

        if facility:
            form_data = {
                "facility": str(facility.pk),
                "medicine": "",
                **cleaned_payload,
            }
            if isinstance(form_data.get("difficulty"), list):
                form_data["difficulty"] = form_data["difficulty"]
            form = FeedbackForm(data=form_data, facility_id=facility.pk)
            if not form.is_valid():
                for field_name, errors in form.errors.items():
                    for error in errors:
                        row_errors.append(
                            {
                                "field_name": field_name,
                                "submitted_value": normalized_row.get(field_name, ""),
                                "error_message": str(error),
                            }
                        )
            cleaned_data = form.cleaned_data if form.is_valid() else {}
        else:
            cleaned_data = {}

        fingerprint = ""
        duplicate_status = DuplicateStatus.NEW
        if facility and ratings:
            fingerprint = build_feedback_fingerprint(
                {
                    "facility_id": facility.id,
                    "submitted_on": submitted_on,
                    "gender": cleaned_payload.get("gender"),
                    "age_group": cleaned_payload.get("age_group"),
                    "distance": cleaned_payload.get("distance"),
                    "service": cleaned_payload.get("service"),
                    "ratings": ratings,
                    "comments": comments,
                }
            )
            if fingerprint in seen_fingerprints:
                duplicate_status = DuplicateStatus.EXACT_DUPLICATE
            else:
                duplicate_status = classify_duplicate(fingerprint, facility.id)
                seen_fingerprints.add(fingerprint)

        status = "valid"
        if row_errors:
            status = "invalid"
            invalid_rows += 1
        elif duplicate_status == DuplicateStatus.EXACT_DUPLICATE:
            status = "exact_duplicate"
            duplicate_rows += 1
        elif duplicate_status == DuplicateStatus.POSSIBLE_DUPLICATE:
            status = "possible_duplicate"
            duplicate_rows += 1
        else:
            valid_rows += 1

        row_results.append(
            {
                "source_row_number": index,
                "status": status,
                "duplicate_status": duplicate_status,
                "errors": row_errors,
                "normalized_row": {key: "" if value is None else str(value) for key, value in normalized_row.items()},
                "facility_id": facility.id if facility else None,
                "submitted_on": submitted_on.isoformat() if submitted_on else "",
                "ratings": ratings,
                "comments": comments,
                "fingerprint": fingerprint,
                "cleaned_payload": cleaned_payload,
            }
        )

    batch.total_rows = len(raw_rows)
    batch.valid_rows = valid_rows
    batch.invalid_rows = invalid_rows
    batch.duplicate_rows = duplicate_rows
    batch.status = ImportBatch.Status.READY if valid_rows else ImportBatch.Status.VALIDATION_FAILED
    batch.validation_summary = {
        "headers": TEMPLATE_COLUMNS,
        "rows": row_results,
        "errors": [],
    }
    batch.save(
        update_fields=[
            "total_rows",
            "valid_rows",
            "invalid_rows",
            "duplicate_rows",
            "status",
            "validation_summary",
        ]
    )
    log_bulk_event("spreadsheet_validated", actor=batch.uploaded_by, import_batch=batch, details={"total_rows": batch.total_rows})
    return batch


def import_validated_batch(batch: ImportBatch, actor):
    if batch.status not in {ImportBatch.Status.READY, ImportBatch.Status.PARTIALLY_COMPLETED}:
        raise ValidationError("Only validated batches can be imported.")

    row_results = batch.validation_summary.get("rows", [])
    imported_rows = 0

    with transaction.atomic():
        batch.status = ImportBatch.Status.IMPORTING
        batch.save(update_fields=["status"])

        for row_result in row_results:
            if row_result["status"] not in {"valid", "possible_duplicate"}:
                continue
            if row_result["status"] == "possible_duplicate" and not ALLOW_POSSIBLE_DUPLICATES:
                continue

            facility = Facility.objects.get(pk=row_result["facility_id"])
            form_data = {
                "facility": str(facility.pk),
                "medicine": "",
                **row_result["cleaned_payload"],
            }
            form = FeedbackForm(data=form_data, facility_id=facility.pk)
            if not form.is_valid():
                continue

            created_entries = create_feedback_entries_from_cleaned_data(
                facility=facility,
                cleaned_data=form.cleaned_data,
                ratings=row_result["ratings"],
                comments=row_result["comments"],
                submission_source=Feedback.SubmissionSource.SPREADSHEET_IMPORT,
                collection_session=batch.collection_session,
                import_batch=batch,
                captured_by=actor,
                submitted_on=datetime.fromisoformat(row_result["submitted_on"]).date(),
                fingerprint=row_result["fingerprint"],
            )
            if created_entries:
                imported_rows += 1

        batch.imported_rows = imported_rows
        batch.completed_at = timezone.now()
        batch.status = (
            ImportBatch.Status.COMPLETED
            if imported_rows == batch.valid_rows
            else ImportBatch.Status.PARTIALLY_COMPLETED
        )
        batch.save(update_fields=["imported_rows", "completed_at", "status"])

    log_bulk_event("import_completed", actor=actor, import_batch=batch, details={"imported_rows": imported_rows})
    return batch


def rollback_import_batch(batch: ImportBatch, actor):
    if batch.rolled_back_at:
        raise ValidationError("This batch has already been rolled back.")
    if batch.status not in {ImportBatch.Status.COMPLETED, ImportBatch.Status.PARTIALLY_COMPLETED}:
        raise ValidationError("Only completed imports can be rolled back.")

    with transaction.atomic():
        affected = batch.feedback_entries.filter(is_active=True)
        if not affected.exists():
            raise ValidationError("No active imported feedback was found for this batch.")
        affected_count = affected.count()
        affected.update(is_active=False, rolled_back_at=timezone.now())
        batch.rolled_back_at = timezone.now()
        batch.rolled_back_by = actor
        batch.status = ImportBatch.Status.ROLLED_BACK
        batch.save(update_fields=["rolled_back_at", "rolled_back_by", "status"])

    log_bulk_event("import_rolled_back", actor=actor, import_batch=batch, details={"affected_count": affected_count})
    return batch


def build_error_report_workbook(batch: ImportBatch):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import Errors"
    sheet.append(
        [
            "source_row_number",
            "row_status",
            "field_name",
            "submitted_value",
            "error_message",
            *TEMPLATE_COLUMNS,
        ]
    )

    for row_result in batch.validation_summary.get("rows", []):
        if row_result["status"] == "valid":
            continue
        errors = row_result.get("errors") or [{"field_name": "", "submitted_value": "", "error_message": ""}]
        for error in errors:
            row_values = row_result.get("normalized_row", {})
            sheet.append(
                [
                    row_result["source_row_number"],
                    row_result["status"],
                    error.get("field_name", ""),
                    error.get("submitted_value", ""),
                    error.get("error_message", ""),
                    *[row_values.get(column, "") for column in TEMPLATE_COLUMNS],
                ]
            )
    return workbook
